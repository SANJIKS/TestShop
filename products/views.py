from rest_framework.views import APIView
from rest_framework.generics import CreateAPIView, ListAPIView
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
from django.http import HttpResponse
from django.core.cache import cache
from .serializers import ProductSerializer
from .models import Product

class ProductView(CreateAPIView, ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ProductSerializer

    def get_queryset(self):
        cached_data = cache.get('products_list')

        if cached_data:
            return cached_data
        
        products = Product.objects.all()
        cache.set('products_list', products, 3600) # Кешируем данные на 1 час

        return products

class ExportProductsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        worksheet = workbook.active

        columns = ['Name', 'Description', 'Price', 'Category', 'Tags', 'Created At']
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        products = Product.objects.select_related('category').prefetch_related('tags').all()
        for row_num, product in enumerate(products, 2):
            worksheet.cell(row=row_num, column=1, value=product.name)
            worksheet.cell(row=row_num, column=2, value=product.description)
            worksheet.cell(row=row_num, column=3, value=product.price)
            worksheet.cell(row=row_num, column=4, value=product.category.name)
            tags = ", ".join(tag.name for tag in product.tags.all())
            worksheet.cell(row=row_num, column=5, value=tags)
            created_at = product.created_at.replace(tzinfo=None)
            worksheet.cell(row=row_num, column=6, value=created_at)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=products.xlsx'
        workbook.save(response)
        return response